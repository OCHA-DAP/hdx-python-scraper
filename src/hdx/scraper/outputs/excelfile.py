import logging
from typing import Dict, List, Optional, Union

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from .base import BaseOutput

try:
    from pandas import DataFrame
except ImportError:
    DataFrame = None


logger = logging.getLogger(__name__)


class ExcelFile(BaseOutput):
    """ExcelFile class enabling writing to Excel spreadsheets.

    Args:
        excel_path (str): Path to output spreadsheet
        tabs (Dict[str, str]): Dictionary of mappings from internal name to spreadsheet tab name
        updatetabs (List[str]): Tabs to update
    """

    def __init__(
        self, excel_path: str, tabs: Dict[str, str], updatetabs: List[str]
    ) -> None:
        super().__init__(updatetabs)
        self.workbook = Workbook()
        self.excel_path = excel_path
        self.tabs = tabs

    def update_tab(
        self,
        tabname: str,
        values: Union[List, DataFrame],
        hxltags: Optional[Dict] = None,
    ) -> None:
        """Update tab with values

        Args:
            tabname (str): Tab to update
            values (Union[List, DataFrame]): Values in a list of lists or a DataFrame
            hxltags (Optional[Dict]): HXL tag mapping. Defaults to None.

        Returns:
            None
        """
        if tabname not in self.updatetabs:
            return
        sheetname = self.tabs[tabname]
        try:
            del self.workbook[sheetname]
        except KeyError:
            pass
        tab = self.workbook.create_sheet(sheetname)
        if isinstance(values, list):
            for i, row in enumerate(values):
                for j, value in enumerate(row):
                    tab.cell(row=i + 1, column=j + 1, value=value)
        else:
            headers = list(values.columns.values)
            tab.append(headers)
            if hxltags:
                tab.append([hxltags.get(header, "") for header in headers])
            for r in dataframe_to_rows(values, index=False, header=False):
                tab.append(r)

    def save(self) -> None:
        """Save spreadsheet

        Returns:
            None
        """
        self.workbook.save(self.excel_path)
