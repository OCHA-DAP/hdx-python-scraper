# -*- coding: utf-8 -*-
import logging
from typing import Dict, List, Union, Optional

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas import DataFrame

logger = logging.getLogger()


class ExcelOutput:
    """ExcelOutput class enabling writing to Excel spreadsheets.

    Args:
        excel_path (str): Path to output spreadsheet
        tabs (Dict[str, str]): Dictionary of mappings from internal name to spreadsheet tab name
        updatetabs (List[str]): Tabs to update
    """

    def __init__(self, excel_path, tabs, updatetabs):
        # type: (str, Dict[str, str], List[str]) -> None
        self.workbook = Workbook()
        self.excel_path = excel_path
        self.tabs = tabs
        self.updatetabs = updatetabs

    def update_tab(self, tabname, values, hxltags=None):
        # type: (str, Union[List, DataFrame], Optional[Dict]) -> None
        """Update tab with values

        Args:
            tabname (str): Tab to update
            values (Union[List, DataFrame]): Either values in a list of dicts or a DataFrame
            hxltags (Optional[Dict]): HXL tag mapping. Defaults to None.

        Returns:
            None
        """
        if tabname not in self.updatetabs:
            return
        sheetname = self.tabs[tabname]
        try:
            del self.workbook[sheetname]
        except KeyError:
            pass
        tab = self.workbook.create_sheet(sheetname)
        if isinstance(values, list):
            for i, row in enumerate(values):
                for j, value in enumerate(row):
                    tab.cell(row=i+1, column=j+1, value=value)
        else:
            headers = list(values.columns.values)
            tab.append(headers)
            if hxltags:
                tab.append([hxltags.get(header, '') for header in headers])
            for r in dataframe_to_rows(values, index=False, header=False):
                tab.append(r)

    def save(self):
        # type: () -> None
        """Save spreadsheet

        Returns:
            None
        """
        self.workbook.save(self.excel_path)
